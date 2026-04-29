"""Build company-level DCF workbook pack for the Indian Equity Valuation Lab.

Output:
outputs/valuation_summary/company_dcf_workbook_pack_v1.xlsx

The workbook contains:
- DCF Summary
- one DCF template/model sheet per company
- Data Dictionary

The workbook is public-source v1 and should be refreshed with audited FY26 data,
verified share counts, net debt/cash reconciliation, beta regression and date-
consistent market data before any final valuation lock.

This script is for education, research practice and portfolio demonstration only.
It is not investment advice.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path("outputs/valuation_summary/company_dcf_workbook_pack_v1.xlsx")

COMPANIES = [
    {"name":"Hero MotoCorp","category":"Core value / quality","sector":"Auto / Two-wheelers","rev":46300,"growths":[8.0,7.0,6.0,5.0,4.5],"margins":[15.0,15.0,15.0,15.0,15.0],"da":1.5,"capex":2.0,"nwc":4.0,"tax":25.2,"wacc":12.58,"tg":4.0,"risk":"Two-wheeler recovery, premium/EV optionality; competition and mature valuation."},
    {"name":"Natco Pharma","category":"Core value / quality","sector":"Pharma / Complex generics","rev":3790,"growths":[-36.7,8.0,10.0,9.0,8.0],"margins":[24.4,26.0,28.0,29.0,29.5],"da":6.5,"capex":8.0,"nwc":5.0,"tax":17.0,"wacc":11.31,"tg":4.0,"risk":"Product-cycle normalization, pipeline optionality, regulatory risk."},
    {"name":"NALCO","category":"Core value / quality","sector":"Metals / Aluminium","rev":17200,"growths":[3.0,4.0,4.0,4.0,3.5],"margins":[36.0,34.0,32.0,30.0,29.0],"da":6.0,"capex":10.0,"nwc":6.0,"tax":25.0,"wacc":15.68,"tg":3.5,"risk":"Mid-cycle aluminium/alumina margin normalization and commodity risk."},
    {"name":"Hindustan Zinc","category":"Core value / quality","sector":"Metals / Zinc-Silver","rev":40000,"growths":[3.0,4.0,4.0,3.5,3.0],"margins":[52.0,50.0,48.0,46.0,45.0],"da":8.0,"capex":10.0,"nwc":5.0,"tax":25.0,"wacc":14.98,"tg":3.5,"risk":"Zinc/silver price cycle, ownership/capital allocation and dividend risk."},
    {"name":"CPCL","category":"Core value / quality","sector":"Oil & Gas / Refining","rev":75000,"growths":[2.0,3.0,3.0,2.5,2.5],"margins":[6.0,5.5,5.0,4.8,4.7],"da":1.5,"capex":3.5,"nwc":5.0,"tax":25.0,"wacc":16.88,"tg":3.0,"risk":"GRM compression, crude inventory losses, working-capital volatility."},
    {"name":"Sharda Cropchem","category":"Core value / quality","sector":"Agrochemicals","rev":4400,"growths":[8.0,7.0,6.0,5.0,4.5],"margins":[17.5,17.0,16.5,16.0,15.5],"da":1.0,"capex":1.5,"nwc":8.0,"tax":25.0,"wacc":14.78,"tg":3.5,"risk":"Working capital, receivables, channel destocking and registration risk."},
    {"name":"Fiem Industries","category":"Growth / satellite","sector":"Auto Components","rev":2750,"growths":[10.0,9.0,8.0,7.0,6.0],"margins":[14.0,13.8,13.6,13.4,13.2],"da":2.5,"capex":4.0,"nwc":6.0,"tax":25.0,"wacc":15.13,"tg":4.0,"risk":"Two-wheeler customer concentration and margin normalization."},
    {"name":"Time Technoplast","category":"Growth / satellite","sector":"Industrial Packaging / Polymers","rev":6000,"growths":[8.0,7.5,7.0,6.0,5.0],"margins":[15.0,15.2,15.3,15.3,15.2],"da":2.5,"capex":4.0,"nwc":7.0,"tax":25.0,"wacc":14.88,"tg":4.0,"risk":"Working capital, raw-material volatility and value-added mix execution."},
    {"name":"Force Motors","category":"Growth / satellite","sector":"Auto / Vans & Engines","rev":8800,"growths":[9.0,8.0,7.0,6.0,5.0],"margins":[15.5,15.0,14.5,14.0,13.5],"da":3.5,"capex":5.0,"nwc":6.0,"tax":25.0,"wacc":15.83,"tg":4.0,"risk":"Exceptional-item normalization, auto-cycle risk and platform concentration."},
    {"name":"Triveni Turbine","category":"Growth / satellite","sector":"Capital Goods / Turbines","rev":2100,"growths":[14.0,12.0,10.0,8.0,7.0],"margins":[23.5,23.0,22.5,22.0,21.5],"da":1.5,"capex":2.5,"nwc":8.0,"tax":25.0,"wacc":14.88,"tg":4.0,"risk":"Order-book conversion, export growth, margin normalization and working capital."},
]

BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
LIGHT_ORANGE = "FCE4D6"
WHITE = "FFFFFF"
BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def style_range(ws, cell_range: str, fill: str | None = None, bold: bool = False, font_color: str | None = None):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if bold or font_color:
                cell.font = Font(bold=bold, color=font_color or "000000")


def autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)


def build_company_sheet(wb: Workbook, company: dict) -> None:
    ws = wb.create_sheet(company["name"][:31])
    ws["A1"] = f"{company['name']} — Public-Source v1 DCF Workbook"
    ws.merge_cells("A1:H1")
    style_range(ws, "A1:H1", BLUE, True, WHITE)
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)

    inputs = [
        ["Input", "Value", "Notes"],
        ["Valuation Status", "Public-Source Complete v1", "Audited-data refresh pending"],
        ["Currency", "INR crore except per-share values", ""],
        ["FY26E Revenue Base", company["rev"], "From company v1 numerical summary"],
        ["D&A % of Revenue", company["da"] / 100, "Public-source v1 assumption"],
        ["Capex % of Revenue", company["capex"] / 100, "Public-source v1 assumption"],
        ["NWC % of Incremental Revenue", company["nwc"] / 100, "Public-source v1 assumption"],
        ["Tax Rate", company["tax"] / 100, "Public-source v1 assumption"],
        ["WACC", company["wacc"] / 100, "Beta/WACC refresh pending"],
        ["Terminal Growth", company["tg"] / 100, "Conservative v1 assumption"],
        ["Net Cash / (Debt)", 0, "Placeholder; refresh from audited balance sheet"],
        ["Shares Outstanding", 0, "Placeholder; refresh from annual report / exchange filing"],
        ["Spot Price", 0, "Placeholder; refresh from NSE/BSE"],
        ["Key Risk", company["risk"], ""],
    ]
    for r, row in enumerate(inputs, 3):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    style_range(ws, "A3:C16")
    style_range(ws, "A3:C3", LIGHT_BLUE, True)
    for cell in ws["B7:B12"]:
        cell[0].number_format = "0.0%"

    ws.append([])
    ws["A4"] = "Revenue Growth"
    ws["A5"] = "EBITDA Margin"
    for col_idx, value in enumerate([None] + [g / 100 for g in company["growths"]], 2):
        ws.cell(4, col_idx, value)
    for col_idx, value in enumerate([None] + [m / 100 for m in company["margins"]], 2):
        ws.cell(5, col_idx, value)
    for row in [4, 5]:
        for col in range(2, 8):
            ws.cell(row, col).number_format = "0.0%"

    start = 19
    years = ["FY26E", "FY27E", "FY28E", "FY29E", "FY30E", "FY31E"]
    metrics = ["Revenue", "Revenue Growth", "EBITDA Margin", "EBITDA", "D&A", "EBIT", "Tax on EBIT", "NOPAT", "Capex", "Change in NWC", "FCFF", "Discount Factor", "PV of FCFF"]
    for c, header in enumerate(["Metric"] + years, 1):
        ws.cell(start, c, header)
    style_range(ws, f"A{start}:G{start}", LIGHT_BLUE, True)
    for i, metric in enumerate(metrics, start + 1):
        ws.cell(i, 1, metric)
    r = {metric: start + 1 + i for i, metric in enumerate(metrics)}
    ws[f"B{r['Revenue']}"] = "=$B$6"
    for col in range(3, 8):
        letter = get_column_letter(col)
        prev = get_column_letter(col - 1)
        ws[f"{letter}{r['Revenue']}"] = f"={prev}{r['Revenue']}*(1+{letter}$4)"
    for col in range(3, 8):
        letter = get_column_letter(col)
        prev = get_column_letter(col - 1)
        ws[f"{letter}{r['Revenue Growth']}"] = f"={letter}{r['Revenue']}/{prev}{r['Revenue']}-1"
        ws[f"{letter}{r['EBITDA Margin']}"] = f"={letter}$5"
    for col in range(2, 8):
        letter = get_column_letter(col)
        ws[f"{letter}{r['EBITDA']}"] = f"={letter}{r['Revenue']}*{letter}{r['EBITDA Margin']}"
        ws[f"{letter}{r['D&A']}"] = f"={letter}{r['Revenue']}*$B$7"
        ws[f"{letter}{r['EBIT']}"] = f"={letter}{r['EBITDA']}-{letter}{r['D&A']}"
        ws[f"{letter}{r['Tax on EBIT']}"] = f"={letter}{r['EBIT']}*$B$10"
        ws[f"{letter}{r['NOPAT']}"] = f"={letter}{r['EBIT']}-{letter}{r['Tax on EBIT']}"
        ws[f"{letter}{r['Capex']}"] = f"={letter}{r['Revenue']}*$B$8"
        if col == 2:
            ws[f"{letter}{r['Change in NWC']}"] = 0
        else:
            prev = get_column_letter(col - 1)
            ws[f"{letter}{r['Change in NWC']}"] = f"=({letter}{r['Revenue']}-{prev}{r['Revenue']})*$B$9"
        ws[f"{letter}{r['FCFF']}"] = f"={letter}{r['NOPAT']}+{letter}{r['D&A']}-{letter}{r['Capex']}-{letter}{r['Change in NWC']}"
    for n, col in enumerate(range(3, 8), 1):
        letter = get_column_letter(col)
        ws[f"{letter}{r['Discount Factor']}"] = f"=1/(1+$B$11)^{n}"
        ws[f"{letter}{r['PV of FCFF']}"] = f"={letter}{r['FCFF']}*{letter}{r['Discount Factor']}"
    style_range(ws, f"A{start}:G{start+len(metrics)}")
    for row in range(r["Revenue Growth"], r["EBITDA Margin"] + 1):
        for col in range(2, 8):
            ws.cell(row, col).number_format = "0.0%"
    for row in range(r["Revenue"], r["PV of FCFF"] + 1):
        for col in range(2, 8):
            if row not in [r["Revenue Growth"], r["EBITDA Margin"], r["Discount Factor"]]:
                ws.cell(row, col).number_format = "#,##0.0"

    outrow = 37
    outputs = [
        ["Valuation Output", "Formula / Value"],
        ["PV of Explicit FCFF", f"=SUM(C{r['PV of FCFF']}:G{r['PV of FCFF']})"],
        ["Terminal Value", f"=G{r['FCFF']}*(1+$B$12)/($B$11-$B$12)"],
        ["PV of Terminal Value", f"=B{outrow+2}*G{r['Discount Factor']}"],
        ["Enterprise Value", f"=B{outrow+1}+B{outrow+3}"],
        ["Net Cash / (Debt)", "=$B$13"],
        ["Equity Value", f"=B{outrow+4}+B{outrow+5}"],
        ["Fair Value / Share", f'=IF($B$14>0,B{outrow+6}/$B$14,"To refresh")'],
        ["Spot Price", '=IF($B$15>0,$B$15,"To refresh")'],
        ["Upside / Downside", f'=IF(AND(ISNUMBER(B{outrow+7}),$B$15>0),B{outrow+7}/$B$15-1,"To refresh")'],
    ]
    for i, row in enumerate(outputs, outrow):
        ws.cell(i, 1, row[0])
        ws.cell(i, 2, row[1])
    style_range(ws, f"A{outrow}:B{outrow+len(outputs)-1}")
    style_range(ws, f"A{outrow}:B{outrow}", LIGHT_BLUE, True)

    srow = 50
    terminal_growths = [0.025, 0.03, 0.035, 0.04, 0.045]
    ws.cell(srow, 1, "WACC \\ g")
    for col, value in enumerate(terminal_growths, 2):
        ws.cell(srow, col, value)
    waccs = [company["wacc"] / 100 - 0.02, company["wacc"] / 100 - 0.01, company["wacc"] / 100, company["wacc"] / 100 + 0.01, company["wacc"] / 100 + 0.02]
    for row_offset, wacc in enumerate(waccs, 1):
        ws.cell(srow + row_offset, 1, wacc)
        for col in range(2, 7):
            letter = get_column_letter(col)
            ws.cell(srow + row_offset, col, f'=IF($A{srow+row_offset}>{letter}${srow},SUM(C{r["PV of FCFF"]}:G{r["PV of FCFF"]})+(G{r["FCFF"]}*(1+{letter}${srow})/($A{srow+row_offset}-{letter}${srow})*G{r["Discount Factor"]}),"n/a")')
    style_range(ws, f"A{srow}:F{srow+5}")
    style_range(ws, f"A{srow}:F{srow}", LIGHT_BLUE, True)

    refresh = [
        ["Refresh dependency", "Status", "Why it matters", "Action", ""],
        ["FY26 audited annual report", "Pending", "Locks audited inputs", "Update revenue/EBIT/capex/NWC", ""],
        ["Share count and net cash/debt", "Pending", "Needed for per-share value", "Reconcile annual report/filing", ""],
        ["Beta regression", "Pending", "Needed for WACC refinement", "Run beta script", ""],
        ["Peer table", "Pending", "Needed for sanity check", "Use date-consistent peer template", ""],
    ]
    for r_i, row in enumerate(refresh, 12):
        for c_i, value in enumerate(row, 4):
            ws.cell(r_i, c_i, value)
    style_range(ws, "D12:H16")
    style_range(ws, "D12:H12", LIGHT_ORANGE, True)

    ws.freeze_panes = "A20"
    autosize(ws)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Summary"
    summary_headers = ["Company", "Category", "Sector", "FY26E Revenue", "FY27 Growth %", "FY27 EBITDA Margin %", "WACC %", "Terminal Growth %", "DCF Sheet", "Refresh Flag"]
    ws.append(summary_headers)
    for c in COMPANIES:
        ws.append([c["name"], c["category"], c["sector"], c["rev"], c["growths"][0], c["margins"][0], c["wacc"], c["tg"], c["name"][:31], "Audited FY26 refresh pending"])
    style_range(ws, f"A1:J{len(COMPANIES)+1}")
    style_range(ws, "A1:J1", BLUE, True, WHITE)
    ws.freeze_panes = "A2"
    autosize(ws)

    for company in COMPANIES:
        build_company_sheet(wb, company)

    dd = wb.create_sheet("Data Dictionary")
    rows = [
        ["Field", "Meaning"],
        ["Public-Source Complete v1", "Research pack complete using public sources; still needs audited FY26 refresh."],
        ["FY26E Revenue Base", "Working revenue base used in preliminary numerical summary."],
        ["EBITDA Margin", "Operating margin assumption before D&A."],
        ["WACC", "Discount rate; beta regression still pending."],
        ["Terminal Growth", "Long-term normalized growth assumption."],
        ["Net Cash / Debt", "Placeholder in workbook until audited balance sheet is reconciled."],
        ["Fair Value / Share", "Calculated only after share count is entered."],
        ["Sensitivity table", "Enterprise value sensitivity to WACC and terminal growth."],
    ]
    for row in rows:
        dd.append(row)
    style_range(dd, f"A1:B{len(rows)}")
    style_range(dd, "A1:B1", BLUE, True, WHITE)
    autosize(dd)
    return wb


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
