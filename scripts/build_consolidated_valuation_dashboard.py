"""Build consolidated valuation dashboard workbook.

This script creates a consolidated Excel-style dashboard for the Indian Equity
Valuation Lab using the current Public-Source Complete v1 company universe.

Output:
outputs/valuation_summary/indian_equity_valuation_dashboard_v1.xlsx

The dashboard is for education, research practice and portfolio demonstration.
It is not investment advice.
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_PATH = Path("outputs/valuation_summary/indian_equity_valuation_dashboard_v1.xlsx")

COMPANIES = [
    ["Hero MotoCorp", "Core value / quality", "Auto / Two-wheelers", "Public-Source Complete v1", "Quality two-wheeler leader with recovery, premiumisation and EV optionality", "Mature business valuation and competition", "FY26 annual report, beta, peer table", "hero_motocorp"],
    ["Natco Pharma", "Core value / quality", "Pharma / Complex generics", "Public-Source Complete v1", "Net-cash niche pharma with pipeline optionality and lumpy earnings", "Product-cycle normalization and regulatory risk", "FY26 audited data, cash/debt, pipeline validation", "natco_pharma"],
    ["NALCO", "Core value / quality", "Metals / Aluminium", "Public-Source Complete v1", "Integrated aluminium/alumina PSU with strong cycle earnings and dividend appeal", "Aluminium/alumina price normalization", "Commodity price deck, FY26 data, peer table", "nalco"],
    ["Hindustan Zinc", "Core value / quality", "Metals / Zinc-Silver", "Public-Source Complete v1", "High-margin zinc/silver producer with dividend relevance", "Zinc/silver price correction and ownership risk", "Zinc/silver deck, FY26 data, peer table", "hindustan_zinc"],
    ["CPCL", "Core value / quality", "Oil & Gas / Refining", "Public-Source Complete v1", "Refining-cycle company with strong GRM recovery and IOCL linkage", "GRM compression and inventory losses", "Mid-cycle GRM deck, FY26 data, debt/cash refresh", "cpcl"],
    ["Sharda Cropchem", "Core value / quality", "Agrochemicals", "Public-Source Complete v1", "Asset-light registration-led agrochemical recovery story", "Working capital, receivables and destocking", "Working-capital days, FY26 data, peer table", "sharda_cropchem"],
    ["Fiem Industries", "Growth / satellite", "Auto Components", "Public-Source Complete v1", "Two-wheeler-linked lighting supplier with LED/EV content optionality", "Customer concentration and margin normalization", "Customer mix, FY26 data, peer table", "fiem_industries"],
    ["Time Technoplast", "Growth / satellite", "Industrial Packaging / Polymers", "Public-Source Complete v1", "Value-added industrial packaging and composite-products growth", "Working capital and raw-material volatility", "Value-added mix, working capital, FY26 data", "time_technoplast"],
    ["Force Motors", "Growth / satellite", "Auto / Vans & Engines", "Public-Source Complete v1", "Auto/OEM and engine manufacturing company with strong operating leverage", "Exceptional-item normalization and auto-cycle risk", "Exceptional item review, FY26 data, segment mix", "force_motors"],
    ["Triveni Turbine", "Growth / satellite", "Capital Goods / Turbines", "Public-Source Complete v1", "High-quality turbine company with export/order-book momentum", "Margin normalization and order-inflow slowdown", "Order book, export/aftermarket mix, FY26 data", "triveni_turbine"],
]

ASSUMPTIONS = [
    ["Hero MotoCorp", 46300, 8.0, 15.0, 12.58, 4.0, "Two-wheeler recovery, premium/EV optionality", "Needs audited FY26 refresh"],
    ["Natco Pharma", 3790, -36.7, 24.4, 11.31, 4.0, "Post-peak pharma normalization and pipeline optionality", "Needs audited FY26 refresh"],
    ["NALCO", 17200, 3.0, 36.0, 15.68, 3.5, "Mid-cycle aluminium/alumina margin normalization", "Needs audited FY26 refresh"],
    ["Hindustan Zinc", 40000, 3.0, 52.0, 14.98, 3.5, "Zinc/silver cycle normalization", "Needs audited FY26 refresh"],
    ["CPCL", 75000, 2.0, 6.0, 16.88, 3.0, "Normalized GRM and refining cycle", "Needs audited FY26 refresh"],
    ["Sharda Cropchem", 4400, 8.0, 17.5, 14.78, 3.5, "Agrochemical recovery with working-capital caution", "Needs audited FY26 refresh"],
    ["Fiem Industries", 2750, 10.0, 14.0, 15.13, 4.0, "Two-wheeler and LED-content growth", "Needs audited FY26 refresh"],
    ["Time Technoplast", 6000, 8.0, 15.0, 14.88, 4.0, "Value-added packaging growth", "Needs audited FY26 refresh"],
    ["Force Motors", 8800, 9.0, 15.5, 15.83, 4.0, "Normalized operating profit excluding exceptional items", "Needs audited FY26 refresh"],
    ["Triveni Turbine", 2100, 14.0, 23.5, 14.88, 4.0, "Order-book and export growth with margin normalization", "Needs audited FY26 refresh"],
]

REFRESH_TASKS = [
    ["FY26 audited annual report reconciliation", "All companies", "High", "Pending", "Update revenue, EBIT, tax, capex, working capital and balance sheet from audited FY26 reports"],
    ["Market price refresh", "All companies", "High", "Pending", "Use one consistent valuation date for price, market cap and EV"],
    ["Share count verification", "All companies", "High", "Pending", "Tie shares outstanding to annual report or exchange filing"],
    ["Net debt / cash reconciliation", "All companies", "High", "Pending", "Reconcile cash, debt, investments and lease liabilities"],
    ["Beta regression", "All companies", "High", "Pending", "Calculate regression beta vs appropriate benchmark index"],
    ["Peer numerical table", "All companies", "High", "Pending", "Collect P/E, EV/EBITDA, margins, ROE, ROCE and leverage from date-consistent source"],
    ["Excel DCF model generation", "Remaining 9 companies", "Medium", "Pending", "Extend Hero workbook generator to all companies"],
    ["Consolidated valuation chart pack", "All companies", "Medium", "Pending", "Create visual dashboard of valuation status and risk themes"],
]

PEERS = [
    ["Hero MotoCorp", "Bajaj Auto; TVS Motor; Eicher Motors; M&M; Ola Electric/Ather reference"],
    ["Natco Pharma", "Divi's Laboratories; Laurus Labs; Granules India; Ajanta Pharma; IPCA Laboratories; Sun Pharma reference"],
    ["NALCO", "Hindalco; Vedanta; Hindustan Zinc; NMDC; Coal India; global aluminium reference"],
    ["Hindustan Zinc", "Vedanta; NALCO; Hindalco; NMDC; Coal India; global zinc/silver reference"],
    ["CPCL", "IOC; BPCL; HPCL; MRPL; Reliance Industries; global refining reference"],
    ["Sharda Cropchem", "UPL; PI Industries; Sumitomo Chemical India; Rallis India; Dhanuka Agritech; Bharat Rasayan"],
    ["Fiem Industries", "Lumax Industries; Minda Corporation; Uno Minda; Varroc Engineering; Endurance Technologies; Suprajit Engineering"],
    ["Time Technoplast", "Mold-Tek Packaging; EPL; Supreme Industries; Nilkamal; TPL Plastech; Hitech Corporation"],
    ["Force Motors", "Ashok Leyland; Tata Motors; M&M; SML Isuzu; Atul Auto; Eicher/VECV context"],
    ["Triveni Turbine", "Thermax; Siemens India; ABB India; Kirloskar Brothers; TD Power Systems; Bharat Bijlee"],
]

LIGHT_BLUE = "EAF3F8"
LIGHT_GREEN = "E2F0D9"
LIGHT_YELLOW = "FFF2CC"
LIGHT_ORANGE = "FCE4D6"
BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def write_rows(ws, rows, start_row=1, start_col=1):
    for r, row in enumerate(rows, start_row):
        for c, value in enumerate(row, start_col):
            cell = ws.cell(r, c, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    coverage = wb.create_sheet("Company Coverage")
    assumptions = wb.create_sheet("DCF Assumptions")
    refresh = wb.create_sheet("Refresh Plan")
    peers = wb.create_sheet("Peer Framework")
    dictionary = wb.create_sheet("Data Dictionary")

    ws["A1"] = "Indian Equity Valuation Lab — Consolidated Dashboard"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:H1")
    write_rows(ws, [
        ["Dashboard status", "Public-Source Complete v1"],
        ["Coverage universe", 10],
        ["Core value / quality", 6],
        ["Growth / satellite", 4],
        ["Completed companies", "=COUNTIF('Company Coverage'!D2:D11,\"Public-Source Complete v1\")"],
        ["Next phase", "Audited FY26 refresh + Excel dashboard automation"],
    ], 3, 1)
    write_rows(ws, [["Metric", "Value", "Interpretation", "Next action"]], 10, 1)
    write_rows(ws, [
        ["Research completion", "=B7/10", "All tracked companies complete at public-source v1 level", "Maintain tracker"],
        ["Audited refresh completion", "0%", "FY26 audited refresh not yet started", "Refresh after annual reports"],
        ["Peer table completion", "0%", "Date-consistent peer numerical tables pending", "Build peer data pull"],
        ["Beta regression completion", "0%", "Regression beta pending for all companies", "Build beta script"],
        ["Excel DCF model completion", "10%", "Hero workbook exists; 9 company workbooks pending", "Extend workbook generator"],
    ], 11, 1)
    write_rows(ws, [["Sector Group", "Companies", "Primary modelling lesson"]], 3, 6)
    write_rows(ws, [
        ["Auto & Components", 3, "Cycle, customer concentration, operating leverage"],
        ["Metals & Commodities", 3, "Mid-cycle price/margin normalization"],
        ["Specialty/Industrial Growth", 4, "Working capital, order book, pipeline/product mix"],
        ["Total", 10, "Normalize strong periods; avoid over-extrapolation"],
        ["Status", "Complete v1", "Data refresh still pending"],
    ], 4, 6)
    style_header(ws, 10)
    for cell in ws[3]:
        cell.font = Font(bold=True)
    ws["B11"].number_format = "0%"
    set_widths(ws, {"A": 24, "B": 30, "C": 46, "D": 40, "F": 26, "G": 16, "H": 48})

    coverage_rows = [["Company", "Category", "Sector / Theme", "Status", "Main Thesis", "Main Risk", "Key Refresh Dependency", "Folder"]] + COMPANIES
    write_rows(coverage, coverage_rows)
    style_header(coverage)
    freeze_and_filter(coverage)
    set_widths(coverage, {"A": 22, "B": 22, "C": 30, "D": 28, "E": 52, "F": 40, "G": 42, "H": 22})
    for row in range(2, 12):
        coverage[f"D{row}"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)

    assumption_rows = [["Company", "FY26E Revenue Base (₹ Cr)", "FY27 Revenue Growth %", "FY27 EBITDA Margin %", "WACC %", "Terminal Growth %", "Base Case Logic", "Refresh Flag"]] + ASSUMPTIONS
    write_rows(assumptions, assumption_rows)
    style_header(assumptions)
    freeze_and_filter(assumptions)
    set_widths(assumptions, {"A": 22, "B": 20, "C": 18, "D": 18, "E": 14, "F": 18, "G": 52, "H": 28})
    for row in range(2, 12):
        assumptions[f"H{row}"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)

    write_rows(refresh, [["Task", "Scope", "Priority", "Status", "Details"]] + REFRESH_TASKS)
    style_header(refresh)
    freeze_and_filter(refresh)
    set_widths(refresh, {"A": 34, "B": 28, "C": 14, "D": 16, "E": 64})
    for row in range(2, 10):
        refresh[f"C{row}"].fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
        refresh[f"D{row}"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)

    write_rows(peers, [["Company", "Peer Set"]] + PEERS)
    style_header(peers)
    freeze_and_filter(peers)
    set_widths(peers, {"A": 24, "B": 90})

    write_rows(dictionary, [["Field", "Meaning", "Notes"],
        ["Public-Source Complete v1", "Research pack is complete using public sources", "Still needs audited financial reconciliation"],
        ["Audited-data refresh pending", "Updated FY26 annual report and filings are required", "Not a missing research section"],
        ["FY26E Revenue Base", "Working revenue base used in v1 numerical summary", "Preliminary until audited data"],
        ["WACC", "Discount rate framework, usually cost of equity dominated", "Beta regression pending"],
        ["Terminal Growth", "Long-term normalized growth assumption", "Should stay below WACC"],
        ["Main Thesis", "Primary reason the company is interesting as a research case", "Not a recommendation"],
        ["Main Risk", "Largest research/valuation risk identified", "Guides sensitivity analysis"],
        ["Peer Framework", "Comparable companies to use as sanity checks", "Numerical peer table pending"],
        ["Refresh Dependency", "Data needed before final valuation lock", "Used for next-phase workflow"],
    ])
    style_header(dictionary)
    freeze_and_filter(dictionary)
    set_widths(dictionary, {"A": 28, "B": 52, "C": 52})

    chart = BarChart()
    chart.title = "Company Count by Group"
    chart.y_axis.title = "Companies"
    chart.x_axis.title = "Group"
    data = Reference(ws, min_col=7, min_row=3, max_row=6)
    cats = Reference(ws, min_col=6, min_row=4, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "A17")

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.value is not None:
                    cell.border = BORDER
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 24

    return wb


if __name__ == "__main__":
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH}")
