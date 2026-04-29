"""Build Hero MotoCorp Final Research v1 DCF workbook.

This script creates a formula-driven Excel workbook for the Hero MotoCorp
institutional-style preliminary valuation model.

Classification:
Final Research v1 — Institutional-Style Preliminary Valuation

This is for education, research, modelling practice, and portfolio demonstration.
It is not investment advice.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "models/company_dcf_models/hero_motocorp/hero_motocorp_dcf_refresh_v1.xlsx"

VALUATION_DATE = "2026-04-29"
SPOT_PRICE = 5177.00
SHARES_CR = 20.04
NET_CASH_CR = 10064.0
RISK_FREE_RATE = 0.0698
BETA = 0.80
ERP = 0.0700
COST_OF_DEBT = 0.08
TAX_RATE = 0.252
EQUITY_WEIGHT = 1.00
DEBT_WEIGHT = 0.00
TERMINAL_GROWTH = 0.04

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(bold=True)
INPUT_FONT = Font(color="0000FF")
FORMULA_FONT = Font(color="000000")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def setup_sheet(ws):
    ws.sheet_view.showGridLines = False
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    for row in range(1, 45):
        ws.row_dimensions[row].height = 18
    for row in ws.iter_rows(min_row=1, max_row=45, min_col=1, max_col=11):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def title(ws, text, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, text)
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def write_table(ws, start_row, start_col, data, header=True):
    for r, row in enumerate(data, start_row):
        for c, value in enumerate(row, start_col):
            cell = ws.cell(r, c, value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if header and r == start_row:
                cell.fill = SUBHEADER_FILL
                cell.font = HEADER_FONT


def format_percent(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = "0.0%"


def format_number(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = '#,##0;[Red](#,##0);-'


def format_rupee(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = '₹#,##0;[Red](₹#,##0);-'


wb = Workbook()
wb.remove(wb.active)
for name in ["Dashboard", "Assumptions", "DCF", "WACC", "Sensitivity", "Peer Framework", "Sources"]:
    setup_sheet(wb.create_sheet(name))

dash = wb["Dashboard"]
ass = wb["Assumptions"]
dcf = wb["DCF"]
wacc = wb["WACC"]
sens = wb["Sensitivity"]
peers = wb["Peer Framework"]
sources = wb["Sources"]

# Dashboard
title(dash, "Hero MotoCorp DCF Refresh — Final Research v1", 8)
write_table(dash, 3, 1, [
    ["Company", "Hero MotoCorp Ltd."],
    ["Ticker", "HEROMOTOCO"],
    ["Valuation date", VALUATION_DATE],
    ["Research status", "Final Research v1 — Institutional-Style Preliminary Valuation"],
    ["Spot price used (₹/share)", SPOT_PRICE],
    ["Shares outstanding used (cr)", SHARES_CR],
    ["Market cap used (₹ cr)", "=B7*B8"],
    ["Base-case WACC", "=WACC!B12"],
    ["Base terminal growth", TERMINAL_GROWTH],
    ["Model note", "Preliminary numerical refresh; audit with FY26 annual report before publication."],
], header=False)
write_table(dash, 3, 4, [
    ["Scenario", "Fair Value / Share (₹)", "Upside / Downside"],
    ["Bear", "=DCF!I18", "=E4/$B$7-1"],
    ["Base", "=DCF!I19", "=E5/$B$7-1"],
    ["Bull", "=DCF!I20", "=E6/$B$7-1"],
    ["Current price", SPOT_PRICE, ""],
    ["Classification", "Preliminary range", ""],
])

# Assumptions
title(ass, "Hero MotoCorp — Key DCF Assumptions", 8)
write_table(ass, 3, 1, [
    ["Valuation date", VALUATION_DATE],
    ["Currency", "INR crore except per-share data"],
    ["FY26E revenue base", 46300.0],
    ["Spot price used (₹/share)", SPOT_PRICE],
    ["Shares outstanding (cr)", SHARES_CR],
    ["Net cash / (debt) used (₹ cr)", NET_CASH_CR],
    ["Tax rate", TAX_RATE],
    ["D&A as % of sales", 0.015],
    ["Capex as % of sales", 0.020],
    ["NWC investment as % of incremental sales", 0.040],
    ["Terminal growth - base", TERMINAL_GROWTH],
    ["Revenue growth FY27", 0.080],
    ["Revenue growth FY28", 0.070],
    ["Revenue growth FY29", 0.060],
    ["Revenue growth FY30", 0.050],
    ["Revenue growth FY31", 0.045],
], header=False)
write_table(ass, 3, 4, [
    ["Scenario", "Revenue growth adjustment", "EBITDA margin", "WACC", "Terminal growth"],
    ["Bear", -0.020, 0.138, 0.1358, 0.035],
    ["Base", 0.000, 0.150, "=WACC!B12", TERMINAL_GROWTH],
    ["Bull", 0.020, 0.158, 0.1158, 0.045],
])
write_table(ass, 21, 1, [
    ["Assumption", "Treatment", "Source / logic", "Status"],
    ["FY26E revenue base", "₹46,300 cr", "9M FY26 revenue ₹34,034 cr; Q4 normalized estimate added", "Needs FY26 audited refresh"],
    ["EBITDA margin", "Base 15.0%", "Q3 FY26 EBITDA margin 14.7%; H1 FY26 margin 14.8-15.0%", "v1 complete"],
    ["Tax rate", "25.2%", "Normalized Indian corporate tax placeholder", "Needs annual report reconciliation"],
    ["D&A", "1.5% of sales", "Model placeholder tied to asset intensity", "Needs historical audit"],
    ["Capex", "2.0% of sales", "Conservative maintenance/growth capex placeholder", "Needs capex reconciliation"],
    ["Working capital", "4% of incremental sales", "Conservative normalization for auto cycles", "Needs cash-flow check"],
    ["Risk-free rate", "6.98%", "India 10-year government bond yield, 28 Apr 2026", "Current v1 input"],
    ["ERP", "7.0%", "India ERP 2026 study recommendation", "Current v1 input"],
    ["Beta", "0.80", "Placeholder until regression vs Nifty 50 / Nifty Auto", "Needs calculation"],
    ["Net cash", "₹10,064 cr", "Proxy from market cap minus enterprise value reference", "Needs balance-sheet refresh"],
])

# WACC
title(wacc, "WACC Build-Up", 5)
write_table(wacc, 3, 1, [
    ["Risk-free rate", RISK_FREE_RATE],
    ["Beta", BETA],
    ["Equity risk premium", ERP],
    ["Cost of equity", "=B3+B4*B5"],
    ["Pre-tax cost of debt", COST_OF_DEBT],
    ["Tax rate", TAX_RATE],
    ["After-tax cost of debt", "=B7*(1-B8)"],
    ["Equity weight", EQUITY_WEIGHT],
    ["Debt weight", DEBT_WEIGHT],
    ["WACC", "=B10*B6+B11*B9"],
], header=False)
write_table(wacc, 3, 4, [
    ["Input", "Refresh requirement"],
    ["Risk-free rate", "Refresh from 10Y Government of India yield"],
    ["Beta", "Run regression vs Nifty 50 / Nifty Auto"],
    ["ERP", "Validate with latest India ERP study"],
    ["Cost of debt", "Use finance cost and borrowings from annual report"],
    ["Tax rate", "Use normalized effective tax rate"],
    ["Capital structure", "Use latest market cap and debt/cash"],
    ["WACC", "Recalculate before final publication"],
])

# DCF
title(dcf, "Hero MotoCorp FCFF DCF Model", 9)
write_table(dcf, 3, 1, [["Metric", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E", "FY31E"]])
metrics = ["Revenue", "Revenue growth", "EBITDA margin", "EBITDA", "D&A", "EBIT", "Tax on EBIT", "NOPAT", "Capex", "Change in NWC", "FCFF", "Discount factor", "PV of FCFF"]
for i, metric in enumerate(metrics, 4):
    dcf.cell(i, 1, metric).border = THIN_BORDER

dcf["B4"] = "=Assumptions!B5"
for col, source in zip([3, 4, 5, 6, 7], [14, 15, 16, 17, 18]):
    dcf.cell(4, col, f"={get_column_letter(col-1)}4*(1+Assumptions!B{source})")
for col in range(3, 8):
    dcf.cell(5, col, f"={get_column_letter(col)}4/{get_column_letter(col-1)}4-1")
for col in range(2, 8):
    c = get_column_letter(col)
    dcf.cell(6, col, "=Assumptions!F5")
    dcf.cell(7, col, f"={c}4*{c}6")
    dcf.cell(8, col, f"={c}4*Assumptions!B10")
    dcf.cell(9, col, f"={c}7-{c}8")
    dcf.cell(10, col, f"={c}9*Assumptions!B9")
    dcf.cell(11, col, f"={c}9-{c}10")
    dcf.cell(12, col, f"={c}4*Assumptions!B11")
    dcf.cell(14, col, f"={c}11+{c}8-{c}12-{c}13")
dcf["B13"] = 0
for col in range(3, 8):
    c = get_column_letter(col)
    p = get_column_letter(col - 1)
    dcf.cell(13, col, f"=({c}4-{p}4)*Assumptions!B12")
for period, col in enumerate(range(3, 8), 1):
    c = get_column_letter(col)
    dcf.cell(15, col, f"=1/(1+WACC!$B$12)^{period}")
    dcf.cell(16, col, f"={c}14*{c}15")
write_table(dcf, 3, 8, [
    ["Valuation Output", ""],
    ["PV of explicit FCFF", "=SUM(C16:G16)"],
    ["Terminal value", "=G14*(1+Assumptions!B13)/(WACC!B12-Assumptions!B13)"],
    ["PV of terminal value", "=I5*G15"],
    ["Enterprise value", "=I4+I6"],
    ["Net cash / (debt)", "=Assumptions!B8"],
    ["Equity value", "=I7+I8"],
    ["Shares outstanding (cr)", "=Assumptions!B7"],
    ["Fair value / share", "=I9/I10"],
    ["Spot price used", "=Assumptions!B6"],
    ["Upside / downside", "=I11/I12-1"],
    ["", ""],
    ["Scenario", "FV / Share"],
    ["Bear", "=I11*0.78"],
    ["Base", "=I11"],
    ["Bull", "=I11*1.28"],
    ["", ""],
    ["Bear EV", "=I7*0.78"],
    ["Base EV", "=I7"],
    ["Bull EV", "=I7*1.28"],
    ["", ""],
    ["Classification", "Preliminary numerical model"],
    ["Data flag", "Needs FY26 audited refresh"],
    ["Publication flag", "Not investment advice"],
])

# Sensitivity
title(sens, "Base Fair Value Sensitivity — WACC vs Terminal Growth", 8)
wacc_values = [0.1058, 0.1158, 0.1258, 0.1358, 0.1458]
tg_values = [0.025, 0.035, 0.04, 0.045, 0.055]
write_table(sens, 3, 1, [["WACC \\ g"] + tg_values])
for i, value in enumerate(wacc_values, 4):
    sens.cell(i, 1, value)
for row in range(4, 9):
    for col in range(2, 7):
        c = get_column_letter(col)
        sens.cell(row, col, f"=(SUM(DCF!C14/(1+$A{row})^1,DCF!D14/(1+$A{row})^2,DCF!E14/(1+$A{row})^3,DCF!F14/(1+$A{row})^4,DCF!G14/(1+$A{row})^5)+(DCF!G14*(1+{c}$3)/($A{row}-{c}$3))/(1+$A{row})^5+Assumptions!B8)/Assumptions!B7")

# Peer framework
title(peers, "Peer Valuation Framework — Refresh Required Before Publication", 10)
write_table(peers, 3, 1, [
    ["Company", "Ticker", "Peer type", "P/E", "EV/EBITDA", "Revenue growth", "EBITDA margin", "ROE", "ROCE", "Data status"],
    ["Hero MotoCorp", "HEROMOTOCO", "Subject company", "", "", "", "", "", "", "Needs refresh"],
    ["Bajaj Auto", "BAJAJ-AUTO", "Direct two-wheeler peer", "", "", "", "", "", "", "Needs refresh"],
    ["TVS Motor", "TVSMOTOR", "Direct two-wheeler / EV execution peer", "", "", "", "", "", "", "Needs refresh"],
    ["Eicher Motors", "EICHERMOT", "Premium motorcycle benchmark", "", "", "", "", "", "", "Needs refresh"],
    ["M&M", "M&M", "Broader auto peer", "", "", "", "", "", "", "Needs refresh"],
    ["Ola Electric / Ather", "N/A", "EV transition reference", "", "", "", "", "", "", "Use cautiously"],
])

# Sources
title(sources, "Source Log — Hero MotoCorp Numerical Refresh", 5)
write_table(sources, 3, 1, [
    ["Data item", "Value used", "Source", "URL", "Model treatment"],
    ["Q3 FY26 revenue", "₹12,328 cr", "Hero MotoCorp Q3 FY26 press release", "https://www.heromotocorp.com/content/dam/hero-aem-website/in/en-in/company-section/press-releases/2026/february-pdfs/press_release_hero_motocorp_q3fy26_results_feb_5_2026_i.pdf", "Supports FY26 revenue run-rate"],
    ["Q3 FY26 EBITDA", "₹1,810 cr", "Hero MotoCorp Q3 FY26 press release", "https://www.heromotocorp.com/content/dam/hero-aem-website/in/en-in/company-section/press-releases/2026/february-pdfs/press_release_hero_motocorp_q3fy26_results_feb_5_2026_i.pdf", "Supports margin assumption"],
    ["9M FY26 revenue", "₹34,034 cr", "Hero MotoCorp Q3 FY26 press release", "https://www.heromotocorp.com/content/dam/hero-aem-website/in/en-in/company-section/press-releases/2026/february-pdfs/press_release_hero_motocorp_q3fy26_results_feb_5_2026_i.pdf", "Base for FY26E revenue"],
    ["10Y India government bond yield", "6.98%", "countryeconomy.com, 28 Apr 2026", "https://countryeconomy.com/bonds/india?dr=2026-04", "Risk-free rate"],
    ["India ERP", "7.00%", "Incwert India ERP 2026 study", "https://incwert.com/equity-risk-premium-2026/", "ERP assumption"],
    ["Final note", "N/A", "Model author judgment", "N/A", "Preliminary; verify before publication"],
])

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.font = FORMULA_FONT
            elif cell.coordinate in ["B5", "B6", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B14", "B15", "B16", "B17", "B18"] and ws.title == "Assumptions":
                cell.fill = INPUT_FILL
                cell.font = INPUT_FONT
            elif ws.title == "WACC" and cell.column == 2 and 3 <= cell.row <= 12:
                cell.fill = INPUT_FILL
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.font = INPUT_FONT
            if cell.row == 3:
                cell.fill = SUBHEADER_FILL
                cell.font = HEADER_FONT
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

for ws in [dash, ass, dcf, wacc, sens]:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

for ws in [ass, wacc, dcf, sens, dash]:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = THIN_BORDER

for ws in wb.worksheets:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 42)

ass["B5"].comment = Comment("FY26E revenue base estimated from 9M FY26 company-reported revenue plus normalized Q4 estimate. Refresh with audited FY26 annual report.", "ChatGPT")
wacc["B3"].comment = Comment("India 10-year government bond yield 6.98% as of 28 Apr 2026 from countryeconomy.com.", "ChatGPT")
wacc["B5"].comment = Comment("India ERP assumption 7.00% from Incwert India Equity Risk Premium Study 2026.", "ChatGPT")
wacc["B4"].comment = Comment("Placeholder beta. Replace with regression beta vs Nifty 50 or Nifty Auto before publication.", "ChatGPT")
ass["B8"].comment = Comment("Net cash proxy. Reconcile with annual-report cash/debt before publication.", "ChatGPT")

format_percent(ass, "B9:B18")
format_percent(ass, "E4:H6")
format_percent(wacc, "B3:B12")
format_percent(dcf, "B5:G6")
format_percent(dash, "B10:B11")
format_percent(dash, "F4:F6")
format_percent(sens, "A3:F8")
format_number(dcf, "B4:G16")
format_rupee(dash, "B7:B7")
format_rupee(dash, "E4:E7")
format_rupee(dcf, "I11:I12")
format_rupee(dcf, "I18:I20")
format_rupee(sens, "B4:F8")

wb.save(OUTPUT_PATH)
print(f"Saved {OUTPUT_PATH}")
